from fileinput import filename
from itertools import product
from openpyxl import load_workbook
from itertools  import product
import pulp

# leer puzzle desde excel
wb = load_workbook("sudoku.xlsx", data_only=True)
sh = wb.worksheets[0]
puzzle = [[[0 for k in range(9)] for j in range(9)] for i in range(9)]
for f,c,v in product(range(9),range(9),range(9)):
    if sh.cell(f+1,c+1).value == v+1:
        puzzle[f][c][v] = 1

#definir prblema
fila = [x for x in range(9)]
colum = [x for x in range(9)]
val = [x for x in range(9)]

prob = pulp.LpProblem("prob", pulp.LpMaximize)

# variables
game = [[[pulp.LpVariable(cat ='Binary', name ='fila {}, col {} es {}'.format(f, c, v+1)) for f in fila] for c in colum] for v in val]

# función objetivo: maximizar el numero de valores asignados
prob += pulp.lpSum(game)

# condición inicial de puzzle
for f,c,v in product(fila,colum,val):
    prob += game[f][c][v] >= puzzle[f][c][v]

# un valor por celda
for f,c in product(fila,colum):
    prob += pulp.lpSum([game[f][c][v] for v in val]) <= 1

# filas valores distintos
for f,v in product(fila,val):
    prob += pulp.lpSum([game[f][c][v] for c in colum]) <= 1

# cols valores distintos
for c,v in product(colum,val):
    prob += pulp.lpSum([game[f][c][v] for f in fila]) <= 1

# cuadrados con valores distintos
for cf,cc,v in product(range(3),range(3),val):
    prob += pulp.lpSum([game[f+3*cf][c+3*cc][v] for f,c in product(range(3),range(3))]) <= 1
    

solver = pulp.getSolver('PULP_CBC_CMD')
prob.solve(solver)

sol = [[0 for k in range(9)] for j in range(9)]
for f,c,v in product(fila,colum,val):
    if(pulp.value(game[f][c][v]) > 0.9):
        sol[f][c] = v + 1

for f in fila:
    print(sol[f])

pulp.value(prob.objective)